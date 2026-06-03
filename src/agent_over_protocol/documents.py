# Copyright (c) 2026 Danny Kim
"""Structured document extraction for workspace files."""

from __future__ import annotations

import asyncio
import csv
from dataclasses import dataclass
from html.parser import HTMLParser
from io import StringIO
from typing import TYPE_CHECKING
from urllib.parse import quote

import httpx
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from collections.abc import Iterable
    from pathlib import Path

type JsonValue = (
    str | int | float | bool | None | list["JsonValue"] | dict[str, "JsonValue"]
)
type JsonObject = dict[str, JsonValue]

TEXT_SUFFIXES = {
    ".cfg",
    ".conf",
    ".css",
    ".csv",
    ".env",
    ".html",
    ".ini",
    ".js",
    ".json",
    ".log",
    ".md",
    ".py",
    ".rst",
    ".text",
    ".toml",
    ".tsv",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}
SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm"}
TIKA_SUFFIXES = {
    ".doc",
    ".docx",
    ".epub",
    ".htm",
    ".odf",
    ".odg",
    ".odp",
    ".ods",
    ".odt",
    ".pdf",
    ".ppt",
    ".pptx",
    ".rtf",
    ".xls",
}
SUPPORTED_SUFFIXES = TEXT_SUFFIXES | SPREADSHEET_SUFFIXES | TIKA_SUFFIXES
HTTP_QUOTED_TEXT_MIN_CODEPOINT = 0x20
HTTP_QUOTED_TEXT_MAX_CODEPOINT = 0x7E


class DocumentReadError(RuntimeError):
    """Raised when a document cannot be extracted."""


@dataclass(frozen=True)
class DocumentReader:
    """Read workspace documents into JSON-serializable structures."""

    tika_url: str
    tika_timeout_seconds: float
    max_spreadsheet_rows: int
    max_spreadsheet_columns: int

    async def read(self, path: Path, *, max_chars: int) -> JsonObject:
        """Read a file as a structured JSON-compatible result."""
        suffix = path.suffix.lower()
        if suffix in SPREADSHEET_SUFFIXES:
            return await asyncio.to_thread(
                self._read_spreadsheet_sync,
                path,
                max_chars,
            )
        if suffix in TEXT_SUFFIXES or not suffix:
            return await asyncio.to_thread(self._read_text_sync, path, max_chars)
        if suffix in TIKA_SUFFIXES:
            return await self._read_with_tika(path, max_chars=max_chars)

        message = f"Unsupported file type: {suffix or 'no extension'}"
        raise DocumentReadError(message)

    async def extract_text(self, path: Path, *, max_chars: int) -> str:
        """Extract searchable plain text from a document."""
        document = await self.read(path, max_chars=max_chars)
        return document_text(document)

    def _read_text_sync(self, path: Path, max_chars: int) -> JsonObject:
        raw_text = _decode_text(path.read_bytes())
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            rows = _delimited_rows(
                raw_text,
                delimiter="\t" if suffix == ".tsv" else ",",
            )
            text = _rows_text(rows)
            return {
                "kind": "table",
                "format": suffix.lstrip(".") or "text",
                "rows": rows,
                "text": _truncate(text, max_chars),
            }
        if suffix in {".html", ".xml"}:
            raw_text = _strip_html(raw_text)
        return {
            "kind": "text_document",
            "format": suffix.lstrip(".") or "text",
            "text": _truncate(raw_text, max_chars),
        }

    def _read_spreadsheet_sync(self, path: Path, max_chars: int) -> JsonObject:
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
            )
        except OSError as exc:
            message = "Spreadsheet could not be read"
            raise DocumentReadError(message) from exc

        sheets: list[JsonValue] = []
        text_parts: list[str] = []
        truncated = False
        for worksheet in workbook.worksheets:
            rows: list[JsonValue] = []
            for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                if row_index > self.max_spreadsheet_rows:
                    truncated = True
                    break
                cells: list[JsonValue] = []
                for column_index, cell in enumerate(row, start=1):
                    if column_index > self.max_spreadsheet_columns:
                        truncated = True
                        break
                    value = _json_cell_value(cell.value)
                    if value is None:
                        continue
                    address = f"{get_column_letter(column_index)}{row_index}"
                    cells.append(
                        {
                            "address": address,
                            "row": row_index,
                            "column": get_column_letter(column_index),
                            "value": value,
                        }
                    )
                if cells:
                    rows.append({"index": row_index, "cells": cells})
                    text_parts.append(_spreadsheet_row_text(worksheet.title, cells))
            sheets.append({"name": worksheet.title, "rows": rows})

        return {
            "kind": "spreadsheet",
            "format": path.suffix.lower().lstrip("."),
            "sheets": sheets,
            "text": _truncate("\n".join(text_parts), max_chars),
            "truncated": truncated,
        }

    async def _read_with_tika(self, path: Path, *, max_chars: int) -> JsonObject:
        data = await asyncio.to_thread(path.read_bytes)
        endpoint = self.tika_url.rstrip("/") + "/rmeta/text"
        try:
            async with httpx.AsyncClient(timeout=self.tika_timeout_seconds) as client:
                response = await client.put(
                    endpoint,
                    content=data,
                    headers={
                        "Accept": "application/json",
                        "Content-Type": "application/octet-stream",
                        "Content-Disposition": _content_disposition(path.name),
                    },
                )
                response.raise_for_status()
        except httpx.HTTPError as exc:
            message = "Tika document extraction failed"
            raise DocumentReadError(message) from exc

        metadata_items = _parse_tika_response(response)
        text = _tika_text(metadata_items)
        metadata = _tika_metadata(metadata_items)
        return {
            "kind": "text_document",
            "format": path.suffix.lower().lstrip("."),
            "metadata": metadata,
            "text": _truncate(text, max_chars),
        }


def is_supported_document(path: Path) -> bool:
    """Return whether the reader can extract text from this path."""
    suffix = path.suffix.lower()
    return suffix in SUPPORTED_SUFFIXES or not suffix


def document_text(document: JsonObject) -> str:
    """Return a plain text view of a structured document result."""
    text = document.get("text")
    if isinstance(text, str):
        return text
    return _flatten_json_text(document)


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp949"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _content_disposition(filename: str) -> str:
    fallback = _ascii_filename_fallback(filename)
    encoded = quote(filename, safe="")
    return f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{encoded}'


def _ascii_filename_fallback(filename: str) -> str:
    fallback = "".join(
        character
        if HTTP_QUOTED_TEXT_MIN_CODEPOINT
        <= ord(character)
        <= HTTP_QUOTED_TEXT_MAX_CODEPOINT
        else "_"
        for character in filename
    )
    fallback = fallback.replace("\\", "\\\\").replace('"', '\\"')
    if fallback.strip(" ._"):
        return fallback
    return "document"


def _delimited_rows(text: str, *, delimiter: str) -> list[JsonValue]:
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows: list[JsonValue] = []
    for index, row in enumerate(reader, start=1):
        rows.append(
            {
                "index": index,
                "cells": [
                    {"index": column_index, "value": value}
                    for column_index, value in enumerate(row, start=1)
                ],
            }
        )
    return rows


def _rows_text(rows: Iterable[JsonValue]) -> str:
    lines: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        cells = row.get("cells")
        if not isinstance(cells, list):
            continue
        values = [
            str(cell.get("value", "")) for cell in cells if isinstance(cell, dict)
        ]
        lines.append(f"{row.get('index')}: {' | '.join(values)}")
    return "\n".join(lines)


def _strip_html(text: str) -> str:
    parser = _TextHTMLParser()
    parser.feed(text)
    return parser.text


class _TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._parts: list[str] = []

    @property
    def text(self) -> str:
        return "\n".join(part for part in self._parts if part)

    def handle_data(self, data: str) -> None:
        stripped = data.strip()
        if stripped:
            self._parts.append(stripped)


def _json_cell_value(value: object) -> JsonValue:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    return str(value)


def _spreadsheet_row_text(sheet_name: str, cells: list[JsonValue]) -> str:
    rendered_cells = [
        f"{cell.get('address')}={cell.get('value')}"
        for cell in cells
        if isinstance(cell, dict)
    ]
    return f"{sheet_name}: {' | '.join(rendered_cells)}"


def _parse_tika_response(response: httpx.Response) -> list[dict[str, JsonValue]]:
    try:
        parsed: object = response.json()
    except ValueError:
        return [{"X-TIKA:content": response.text}]
    if isinstance(parsed, list):
        return [_json_object(item) for item in parsed if isinstance(item, dict)]
    if isinstance(parsed, dict):
        return [_json_object(parsed)]
    return [{"X-TIKA:content": str(parsed)}]


def _json_object(value: object) -> dict[str, JsonValue]:
    if not isinstance(value, dict):
        return {}
    return {str(key): _json_value(item) for key, item in value.items()}


def _json_value(value: object) -> JsonValue:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return _json_object(value)
    return str(value)


def _tika_text(items: list[dict[str, JsonValue]]) -> str:
    parts: list[str] = []
    for item in items:
        content = item.get("X-TIKA:content")
        if isinstance(content, str) and content.strip():
            parts.append(content.strip())
    return "\n\n".join(parts)


def _tika_metadata(items: list[dict[str, JsonValue]]) -> JsonObject:
    metadata: JsonObject = {}
    for item in items:
        for key, value in item.items():
            if key != "X-TIKA:content":
                metadata[key] = value
    return metadata


def _flatten_json_text(value: JsonValue) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, int | float | bool):
        return str(value)
    if isinstance(value, list):
        return "\n".join(_flatten_json_text(item) for item in value)
    return "\n".join(_flatten_json_text(item) for item in value.values())


def _truncate(content: str, max_chars: int) -> str:
    stripped = content.strip()
    if max_chars <= 0 or len(stripped) <= max_chars:
        return stripped
    return (
        f"{stripped[:max_chars].rstrip()}\n\n"
        f"[Document truncated to {max_chars} characters.]"
    )
